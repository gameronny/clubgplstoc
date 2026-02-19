import tkinter as tk
from tkinter import ttk
import database
from PIL import Image, ImageTk
import pandas as pd
from tkinter import filedialog
from tkcalendar import DateEntry

# ---------------------
# Culori (constante)
# ---------------------
MAIN_BG = "#2C3E50"     # fundal fereastra principală
FRAME_BG = "#34495E"    # culoare frame "Adaugă produs" -> folosită pentru dialoguri
TEXT_COLOR = "white"
GREEN = "#2ECC71"       # butoane confirmare
RED = "#E74C3C"         # butoane anulare/ștergere
WARNING = "#E67E22"     # avertizare (necesar pentru label-uri)
HEADER_BG = GREEN       # header Treeview (am folosit verde pentru accent)

# === Funcție confirmare Da/Nu ===
def confirm_dialog(parent, title, message):
    dialog = tk.Toplevel(parent)
    dialog.title(title)

    width, height = 400, 150
    parent.update_idletasks()
    x = parent.winfo_rootx() + (parent.winfo_width() // 2) - (width // 2)
    y = parent.winfo_rooty() + (parent.winfo_height() // 2) - (height // 2)
    dialog.geometry(f"{width}x{height}+{x}+{y}")

    # fundal = culoarea frame "Adaugă produs"
    dialog.configure(bg=FRAME_BG)
    dialog.grab_set()

    tk.Label(dialog, text=message, font=("Arial", 10, "bold"), bg=FRAME_BG, fg=TEXT_COLOR).pack(pady=20)

    result = {"value": False}

    def confirm():
        result["value"] = True
        dialog.destroy()

    def cancel():
        result["value"] = False
        dialog.destroy()

    btn_frame = tk.Frame(dialog, bg=FRAME_BG)
    btn_frame.pack(pady=10)
    tk.Button(btn_frame, text="Da", command=confirm, bg=GREEN, fg=TEXT_COLOR, width=10).pack(side="left", padx=10)
    tk.Button(btn_frame, text="Nu", command=cancel, bg=RED, fg=TEXT_COLOR, width=10).pack(side="right", padx=10)

    dialog.wait_window()
    return result["value"]


# === Fereastra Stock ===
class StockWindow:
    def __init__(self, master):
        self.top = tk.Toplevel(master)
        self.top.title("Gestionează Stocul")
        self.top.geometry("1100x650")
        self.top.state('zoomed')

        # --- FIX focus + stacking ---
          # leagă de fereastra principală
        self.top.focus_force()  # pune focus pe StockWindow
        self.top.lift()  # aduce fereastra în față

        # culori folosite în instanță
        self.bg_color = MAIN_BG
        self.frame_bg_color = FRAME_BG
        self.text_color = TEXT_COLOR
        self.accent_color = GREEN
        self.error_color = RED
        self.warning_color = WARNING

        self.top.configure(bg=self.bg_color)

        # --- Căutare produs ---
        search_frame = tk.Frame(self.top, bg=self.bg_color)
        search_frame.pack(pady=10)
        self.entry_search = tk.Entry(search_frame, width=30, bg=self.frame_bg_color, fg=self.text_color, insertbackground=self.text_color)
        self.entry_search.pack(side="left", padx=5)
        tk.Button(search_frame, text="Caută", command=self.search_product, bg=self.accent_color, fg=TEXT_COLOR, width=12).pack(side="left", padx=5)
        tk.Button(search_frame, text="Resetează", command=self.load_products, bg=self.error_color, fg=TEXT_COLOR, width=12).pack(side="left", padx=5)

        # --- Frame Adaugă / Actualizează ---
        top_frame = tk.Frame(self.top, bg=self.bg_color)
        top_frame.pack(fill="x", pady=10, padx=10)

        # --- FRAME 3D pentru "Adaugă produs" ---
        shadow_add = tk.Frame(top_frame, bg=self.frame_bg_color)
        shadow_add.pack(side="left", padx=(0, 15), pady=5, ipadx=2, ipady=2)

        add_frame = tk.LabelFrame(
            shadow_add,
            text="Adaugă produs",
            bg=self.bg_color,
            fg=self.text_color,
            font=("Arial", 10, "bold"),
            width=300, height=200,
            bd=0, highlightthickness=0, relief="flat",
            padx=15, pady=15
        )
        add_frame.pack()

        tk.Label(add_frame, text="Nume:", bg=self.bg_color, fg=self.text_color).grid(row=0, column=0, sticky="e", pady=5, padx=5)
        self.entry_add_name = tk.Entry(add_frame, width=30, bg=self.bg_color, fg=self.text_color, insertbackground=self.text_color)
        self.entry_add_name.grid(row=0, column=1, pady=5, padx=5)

        tk.Label(add_frame, text="Preț:", bg=self.bg_color, fg=self.text_color).grid(row=1, column=0, sticky="e", pady=5, padx=5)
        self.entry_add_price = tk.Entry(add_frame, width=30, bg=self.bg_color, fg=self.text_color, insertbackground=self.text_color)
        self.entry_add_price.grid(row=1, column=1, pady=5, padx=5)

        tk.Label(add_frame, text="Cantitate:", bg=self.bg_color, fg=self.text_color).grid(row=2, column=0, sticky="e", pady=5, padx=5)
        self.entry_add_qty = tk.Entry(add_frame, width=30, bg=self.bg_color, fg=self.text_color, insertbackground=self.text_color)
        self.entry_add_qty.grid(row=2, column=1, pady=5, padx=5)

        tk.Button(add_frame, text="Adaugă", bg=self.accent_color, fg=TEXT_COLOR, width=25, command=self.adauga_produs).grid(row=3, column=0, columnspan=2, pady=10)

        # --- FRAME LOGO CENTRAT ---
        logo_frame = tk.Frame(top_frame, bg=self.bg_color)
        logo_frame.pack(side="left", fill="both", expand=True)

        import os
        import sys
        from PIL import Image, ImageTk

        try:
            # compatibil și în modul PyInstaller (.exe)
            if getattr(sys, 'frozen', False):
                base_path = sys._MEIPASS
            else:
                base_path = os.path.dirname(os.path.abspath(__file__))

            logo_path = os.path.join(base_path, "logo.png")

            image = Image.open(logo_path)
            image = image.resize((400, 150), Image.LANCZOS)
            self.logo_img = ImageTk.PhotoImage(image)
            self.logo_label = tk.Label(logo_frame, image=self.logo_img, bg=self.bg_color)
            self.logo_label.pack(expand=True)

        except Exception as e:
            print(f"Eroare la încărcarea logo-ului în StockWindow: {e}")
            self.logo_label = tk.Label(logo_frame, text="Logo lipsește", bg=self.error_color, fg=self.text_color)
            self.logo_label.pack(expand=True)

        # --- FRAME 3D pentru "Actualizează produs" ---
        shadow_update = tk.Frame(top_frame, bg=self.frame_bg_color)
        shadow_update.pack(side="right", padx=(15, 0), pady=5, ipadx=2, ipady=2)

        update_frame = tk.LabelFrame(
            shadow_update,
            text="Actualizează produs",
            bg=self.bg_color,
            fg=self.text_color,
            font=("Arial", 10, "bold"),
            width=400, height=200,
            bd=0, highlightthickness=0, relief="flat",
            padx=10, pady=10
        )
        update_frame.pack()

        tk.Label(update_frame, text="ID Produs:", bg=self.bg_color, fg=self.text_color).grid(row=0, column=0, sticky="e", pady=5, padx=5)
        self.entry_update_id = tk.Entry(update_frame, width=30, bg=self.bg_color, fg=self.text_color, insertbackground=self.text_color)
        self.entry_update_id.grid(row=0, column=1, pady=5, padx=5)

        tk.Label(update_frame, text="Nume nou:", bg=self.bg_color, fg=self.text_color).grid(row=1, column=0, sticky="e", pady=5, padx=5)
        self.entry_update_name = tk.Entry(update_frame, width=30, bg=self.bg_color, fg=self.text_color, insertbackground=self.text_color)
        self.entry_update_name.grid(row=1, column=1, pady=5, padx=5)

        tk.Label(update_frame, text="Preț nou:", bg=self.bg_color, fg=self.text_color).grid(row=2, column=0, sticky="e", pady=5, padx=5)
        self.entry_update_price = tk.Entry(update_frame, width=30, bg=self.bg_color, fg=self.text_color, insertbackground=self.text_color)
        self.entry_update_price.grid(row=2, column=1, pady=5, padx=5)

        tk.Label(update_frame, text="Cantitate nouă:", bg=self.bg_color, fg=self.text_color).grid(row=3, column=0, sticky="e", pady=5, padx=5)
        self.entry_update_qty = tk.Entry(update_frame, width=30, bg=self.bg_color, fg=self.text_color, insertbackground=self.text_color)
        self.entry_update_qty.grid(row=3, column=1, pady=5, padx=5)

        tk.Button(update_frame, text="Actualizează", bg=self.accent_color, fg=TEXT_COLOR, width=25, command=self.actualizeaza_produs).grid(
            row=4, column=0, columnspan=2, pady=10)

        # --- Avertizare și confirmări ---
        self.alert_label = tk.Label(self.top, text="", fg=self.warning_color, bg=self.bg_color, font=("Arial", 12, "bold"))
        self.alert_label.pack(pady=5)

        self.confirm_label = tk.Label(self.top, text="", fg=self.accent_color, bg=self.bg_color, font=("Arial", 12, "bold"))
        self.confirm_label.pack(pady=5)

        # --- Tabel produse ---
        columns = ("ID", "Nume", "Preț", "Cantitate", "Vândute", "Încasate", "Neîncasate")
        tree_frame = tk.Frame(self.top, bg=self.bg_color)
        tree_frame.pack(fill="both", expand=True, pady=10, padx=10)

        self.tree = ttk.Treeview(tree_frame, columns=columns, show="headings")
        self.tree.pack(fill="both", expand=True, side="left")

        scrollbar = ttk.Scrollbar(tree_frame, orient="vertical", command=self.tree.yview)
        scrollbar.pack(side="right", fill="y")
        self.tree.configure(yscrollcommand=scrollbar.set)

        for col in columns:
            self.tree.heading(col, text=col)
            self.tree.column(col, width=130, minwidth=50, stretch=True)

        self.tree.column("ID", anchor="e", width=50)
        self.tree.column("Nume", anchor="e", width=150)
        self.tree.column("Preț", anchor="e", width=100)
        self.tree.column("Cantitate", anchor="e", width=100)
        self.tree.column("Vândute", anchor="e", width=100)
        self.tree.column("Încasate", anchor="e", width=100)
        self.tree.column("Neîncasate", anchor="e", width=100)

        self.tree.tag_configure('evenrow', background='white', foreground='black')
        self.tree.tag_configure('oddrow', background='#f0f0f0', foreground='black')
        self.tree.tag_configure('low_stock', background='#f8d7da', foreground='black')  # roșu deschis

        style = ttk.Style()
        style.theme_use("default")
        style.configure("Treeview.Heading",
                        background="#566573",  # culoarea nouă
                        foreground=self.text_color,  # sau TEXT_COLOR
                        font=("Arial", 10, "bold"))

        # --- Butoane acțiuni ---
        btn_frame = tk.Frame(self.top, bg=self.bg_color)
        btn_frame.pack(pady=15)
        tk.Button(btn_frame, text="Vinde", command=self.vinde, bg=self.error_color, fg=TEXT_COLOR, width=20).pack(side="left", padx=10)
        tk.Button(btn_frame, text="Încasează", command=self.incaseaza, bg=self.accent_color, fg=TEXT_COLOR, width=20).pack(side="left", padx=20)
        tk.Button(btn_frame, text="Șterge", command=self.sterge_produs, bg=self.error_color, fg=TEXT_COLOR, width=20).pack(side="left", padx=20)
        tk.Button(btn_frame, text="Raport", bg="#FFA500", fg=TEXT_COLOR, width=20, command=self.open_raport).pack(
            side="right", padx=10)
        self.load_products()

    def clear_entries(self, entries):
        """Șterge conținutul câmpurilor Entry din lista `entries`."""
        for entry in entries:
            entry.delete(0, tk.END)

    def show_message(self, text, color, timeout=5000):
        self.confirm_label.config(text=text, fg=color)
        self.top.after(timeout, lambda: self.confirm_label.config(text=""))

    # === Metode Adaugă / Actualizează ===

    def adauga_produs(self):
        name = self.entry_add_name.get().strip()
        try:
            price = float(self.entry_add_price.get())
            qty = int(self.entry_add_qty.get())
        except ValueError:
            self.confirm_label.config(text="Preț și Cantitate trebuie să fie numere.", fg=self.error_color)
            return

        if not name:
            self.confirm_label.config(text="Numele produsului nu poate fi gol.", fg=self.error_color)
            return

        conn = database.get_connection()
        cursor = conn.cursor()

        # verificăm dacă produsul există deja (ignorând majusculele)
        cursor.execute("SELECT IdProduct, Pret, Cantitate FROM Products WHERE LOWER(Nume) = LOWER(?)", (name,))
        existing = cursor.fetchone()

        unit = "bucată" if qty == 1 else "bucăți"

        if existing:
            # produsul există -> restoc
            prod_id, old_price, old_qty = existing
            new_qty = old_qty + qty

            # dacă vrei să păstreze prețul vechi, comentează linia de mai jos
            cursor.execute("UPDATE Products SET Cantitate = ? WHERE IdProduct = ?", (new_qty, prod_id))
            msg = f"S-au restocat {qty} {unit} '{name}'. Total: {new_qty}."

        else:
            # produs nou -> adăugare
            cursor.execute(
                "INSERT INTO Products (Nume, Pret, Cantitate, Vandute, Incasate, Neincasate) VALUES (?, ?, ?, 0, 0, 0)",
                (name, price, qty))
            msg = f"S-au adăugat {qty} {unit} '{name}'."

        conn.commit()
        conn.close()

        # actualizăm tabelul după modificarea bazei
        self.load_products()

        # afișăm mesajul
        self.confirm_label.config(text=msg, fg=self.accent_color)
        self.top.after(8000, lambda: self.confirm_label.config(text=""))  # dispare după 3 secunde

        # curățăm câmpurile
        self.clear_entries([self.entry_add_name, self.entry_add_price, self.entry_add_qty])

    def actualizeaza_produs(self):
        try:
            prod_id = int(self.entry_update_id.get())
            new_name = self.entry_update_name.get().strip()
            new_price = float(self.entry_update_price.get())
            new_qty = int(self.entry_update_qty.get())
        except ValueError:
            self.confirm_label.config(text="ID, Preț și Cantitate trebuie să fie numere.", fg=self.error_color)
            return
        if not new_name:
            self.confirm_label.config(text="Numele nu poate fi gol.", fg=self.error_color)
            return

        conn = database.get_connection()
        cursor = conn.cursor()
        cursor.execute("""
            UPDATE Products
            SET Nume = ?, Pret = ?, Cantitate = ?
            WHERE IdProduct = ?
        """, (new_name, new_price, new_qty, prod_id))
        conn.commit()
        conn.close()

        self.load_products()
        self.confirm_label.config(text=f"Produsul ID {prod_id} a fost actualizat.", fg=self.accent_color)
        self.clear_entries(
            [self.entry_update_id, self.entry_update_name, self.entry_update_price, self.entry_update_qty])

    def load_products(self):

        self.entry_search.delete(0, tk.END)
        for i in self.tree.get_children():
            self.tree.delete(i)

        conn = database.get_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM Products")
        rows = cursor.fetchall()
        conn.close()

        for idx, row in enumerate(rows):
            incasate_cant = int(row[5] // row[2]) if row[2] else 0
            neincasate_cant = int(row[6] // row[2]) if row[2] else 0
            display_row = (row[0], row[1], row[2], row[3], row[4], incasate_cant, neincasate_cant)

            # Evidențiere stoc redus
            if row[3] < 5:  # prag: 5 bucăți
                tag = 'low_stock'
            else:
                tag = 'evenrow' if idx % 2 == 0 else 'oddrow'

            self.tree.insert("", tk.END, values=display_row, tags=(tag,))

        low_stock_products = [row[1] for row in rows if row[3] < 5]
        self.alert_label.config(text="Stoc redus: " + ", ".join(low_stock_products) if low_stock_products else "", fg=self.warning_color)
        self.confirm_label.config(text="")

    def search_product(self):
        query = self.entry_search.get().strip()
        conn = database.get_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM Products WHERE Nume LIKE ?", ('%' + query + '%',))
        rows = cursor.fetchall()
        conn.close()

        for i in self.tree.get_children():
            self.tree.delete(i)

        for idx, row in enumerate(rows):
            incasate_cant = int(row[5] // row[2]) if row[2] else 0
            neincasate_cant = int(row[6] // row[2]) if row[2] else 0
            display_row = (row[0], row[1], row[2], row[3], row[4], incasate_cant, neincasate_cant)

            # Evidențiere stoc redus
            if row[3] < 5:
                tag = 'low_stock'
            else:
                tag = 'evenrow' if idx % 2 == 0 else 'oddrow'

            self.tree.insert("", tk.END, values=display_row, tags=(tag,))

        low_stock_products = [row[1] for row in rows if row[3] < 5]
        self.alert_label.config(text="Stoc redus: " + ", ".join(low_stock_products) if low_stock_products else "", fg=self.warning_color)
        self.confirm_label.config(text="")

    def vinde(self):
        selected_item = self.tree.selection()
        if not selected_item:
            self.confirm_label.config(text="Selectează un produs pentru vânzare.", fg=self.error_color)
            return
        item_details = self.tree.item(selected_item)['values']
        product_id = item_details[0]
        product_name = item_details[1]

        qty_window = tk.Toplevel(self.top)
        qty_window.title(f"Vinde - '{product_name}'")

        width, height = 400, 150
        self.top.update_idletasks()
        x = self.top.winfo_rootx() + (self.top.winfo_width() // 2) - (width // 2)
        y = self.top.winfo_rooty() + (self.top.winfo_height() // 2) - (height // 2)
        qty_window.geometry(f"{width}x{height}+{x}+{y}")

        # fundal = culoarea frame "Adaugă produs"
        qty_window.configure(bg=self.frame_bg_color)
        qty_window.grab_set()

        tk.Label(qty_window, text=f"Câte bucăți din '{product_name}' vrei să vinzi?", font=("Arial", 10, "bold"),
                 bg=self.frame_bg_color, fg=self.text_color).pack(pady=20)
        qty_entry = tk.Entry(qty_window)
        qty_entry.pack(pady=5)

        def confirm_vanzare():
            try:
                qty = int(qty_entry.get())
                if qty <= 0:
                    raise ValueError

                # Verifică dacă există suficient stoc
                if qty > item_details[3]:  # item_details[3] = Cantitate disponibilă
                    self.confirm_label.config(text="Cantitate insuficientă în stoc.", fg=self.error_color)
                    return

                conn = database.get_connection()
                cursor = conn.cursor()
                cursor.execute("""
                    UPDATE Products
                    SET Cantitate = Cantitate - ?, Vandute = Vandute + ?, Neincasate = Neincasate + (? * Pret), DataVanzare = DATE('now')
                    WHERE IdProduct = ?
                """, (qty, qty, qty, product_id))
                conn.commit()
                conn.close()

                # Curățare câmp Entry înainte de a distruge fereastra
                self.clear_entries([qty_entry])

                # Distruge fereastra
                qty_window.destroy()
                self.load_products()

                # Mesaj confirmare
                unit = "bucată" if qty == 1 else "bucăți"
                self.confirm_label.config(
                    text=f"S-a vândut {qty} {unit} din '{product_name}'." if qty == 1 else f"Au fost vândute {qty} {unit} din '{product_name}'.",
                    fg=self.accent_color
                )

            except ValueError:
                self.confirm_label.config(text="Introdu o cantitate validă.", fg=self.error_color)

        btn_frame = tk.Frame(qty_window, bg=self.frame_bg_color)
        btn_frame.pack(pady=15)
        tk.Button(btn_frame, text="Confirmă", command=confirm_vanzare, bg=self.accent_color, fg=TEXT_COLOR, width=12).pack(side="left", padx=10)
        tk.Button(btn_frame, text="Anulează", command=qty_window.destroy, bg=self.error_color, fg=TEXT_COLOR, width=12).pack(side="left", padx=10)

    def incaseaza(self):
        selected_item = self.tree.selection()
        if not selected_item:
            self.show_message("Selectează un produs pentru încasare.", self.error_color)
            return
        item_details = self.tree.item(selected_item)['values']
        product_id = item_details[0]
        product_name = item_details[1]
        neincasate = int(item_details[6])
        if neincasate == 0:
            self.show_message(f"Produsul '{product_name}' nu are sume neîncasate.", self.error_color)
            return

        qty_window = tk.Toplevel(self.top)
        qty_window.title(f"Încasează - {product_name}")

        width, height = 400, 150
        self.top.update_idletasks()
        x = self.top.winfo_rootx() + (self.top.winfo_width() // 2) - (width // 2)
        y = self.top.winfo_rooty() + (self.top.winfo_height() // 2) - (height // 2)
        qty_window.geometry(f"{width}x{height}+{x}+{y}")
        qty_window.configure(bg=self.frame_bg_color)
        qty_window.grab_set()

        tk.Label(qty_window, text=f"Câte bucăți din '{product_name}' vrei să încasezi?", font=("Arial", 10, "bold"),
                 bg=self.frame_bg_color, fg=self.text_color).pack(pady=20)
        qty_entry = tk.Entry(qty_window)
        qty_entry.pack(pady=5)

        def confirm_incasare():
            try:
                qty = int(qty_entry.get())
                if qty <= 0 or qty > neincasate:
                    raise ValueError

                conn = database.get_connection()
                cursor = conn.cursor()
                cursor.execute("""
                    UPDATE Products
                    SET Incasate = Incasate + (? * Pret), Neincasate = Neincasate - (? * Pret)
                    WHERE IdProduct = ?
                """, (qty, qty, product_id))
                conn.commit()
                conn.close()

                # Curățare câmp înainte de a distruge fereastra
                self.clear_entries([qty_entry])

                # distruge fereastra
                qty_window.destroy()
                self.load_products()
                unit = "bucată" if qty == 1 else "bucăți"
                self.confirm_label.config(
                    text=f"S-a încasat {qty} {unit} din '{product_name}'." if qty == 1 else f"S-au încasat {qty} {unit} din '{product_name}'.",
                    fg=self.accent_color)

            except ValueError:
                self.confirm_label.config(text=f"Introdu o cantitate validă (1-{neincasate}).", fg=self.error_color)

        btn_frame = tk.Frame(qty_window, bg=self.frame_bg_color)
        btn_frame.pack(pady=15)
        tk.Button(btn_frame, text="Confirmă", command=confirm_incasare, bg=self.accent_color, fg=self.text_color,
                  width=12).pack(side="left", padx=10)
        tk.Button(btn_frame, text="Anulează", command=qty_window.destroy, bg=self.error_color, fg=self.text_color,
                  width=12).pack(side="left", padx=10)

    def sterge_produs(self):
        selected_item = self.tree.selection()
        if not selected_item:
            self.confirm_label.config(text="Selectează un produs pentru ștergere.", fg=self.error_color)
            return
        item_details = self.tree.item(selected_item)['values']
        product_id = item_details[0]
        product_name = item_details[1]

        answer = confirm_dialog(self.top, "Confirmare", f"Ești sigur că vrei să ștergi produsul '{product_name}'?")
        if not answer:
            return

        conn = database.get_connection()
        cursor = conn.cursor()
        cursor.execute("DELETE FROM Products WHERE IdProduct = ?", (product_id,))
        conn.commit()
        conn.close()

        self.load_products()
        self.confirm_label.config(text=f"Produsul '{product_name}' a fost șters.", fg=self.accent_color)

    def open_raport(self):
        from tkinter import filedialog, messagebox
        from tkcalendar import DateEntry
        import pandas as pd
        from openpyxl import load_workbook
        from openpyxl.styles import Alignment

        raport_window = tk.Toplevel(self.top)
        raport_window.title("Generează raport")
        raport_window.configure(bg=self.frame_bg_color)

        # poziționare centrată față de self.top
        width, height = 400, 250
        self.top.update_idletasks()
        x = self.top.winfo_rootx() + (self.top.winfo_width() // 2) - (width // 2)
        y = self.top.winfo_rooty() + (self.top.winfo_height() // 2) - (height // 2)
        raport_window.geometry(f"{width}x{height}+{x}+{y}")

        # permit închiderea cu X normal
        raport_window.protocol("WM_DELETE_WINDOW", raport_window.destroy)

        # câmpuri DateEntry
        tk.Label(raport_window, text="Data start:", bg=self.frame_bg_color, fg=self.text_color).pack(pady=8)
        start_date = DateEntry(raport_window, width=20, background="darkblue", foreground="white",
                               borderwidth=2, date_pattern='dd/mm/yyyy')
        start_date.pack(pady=2)

        tk.Label(raport_window, text="Data sfârșit:", bg=self.frame_bg_color, fg=self.text_color).pack(pady=8)
        end_date = DateEntry(raport_window, width=20, background="darkblue", foreground="white",
                             borderwidth=2, date_pattern='dd/mm/yyyy')
        end_date.pack(pady=2)

        def genereaza_raport():
            start = start_date.get_date().strftime("%Y-%m-%d")
            end = end_date.get_date().strftime("%Y-%m-%d")

            conn = database.get_connection()
            cursor = conn.cursor()
            cursor.execute("""
                SELECT * FROM Products
                WHERE DataVanzare BETWEEN ? AND ?
            """, (start, end))
            rows = cursor.fetchall()
            conn.close()

            if not rows:
                messagebox.showinfo("Raport", "Nu există date pentru perioada selectată.", parent=raport_window)
                return

            df = pd.DataFrame(rows, columns=[
                "ID", "Nume", "Pret", "Cantitate", "Vandute",
                "Incasate", "Neincasate", "DataVanzare"
            ])
            df['DataVanzare'] = pd.to_datetime(df['DataVanzare']).dt.strftime("%d/%m/%Y")
            df['Incasate'] = (pd.to_numeric(df['Incasate'], errors='coerce') / pd.to_numeric(df['Pret'],
                                                                                             errors='coerce')).fillna(
                0).astype(int)
            df['Neincasate'] = (pd.to_numeric(df['Neincasate'], errors='coerce') / pd.to_numeric(df['Pret'],
                                                                                                 errors='coerce')).fillna(
                0).astype(int)

            # dialogul Save As
            file_path = filedialog.asksaveasfilename(parent=raport_window, defaultextension=".xlsx",
                                                     filetypes=[("Excel files", "*.xlsx")])
            if not file_path:
                return  # dacă apasă Cancel, fereastra rămâne deschisă

            df.to_excel(file_path, index=False)

            # aliniere la dreapta cu openpyxl
            wb = load_workbook(file_path)
            ws = wb.active
            for r in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in r:
                    cell.alignment = Alignment(horizontal='right')
            wb.save(file_path)

            # confirmare și închidere imediată
            show_custom_message(raport_window, "Raport salvat", f"Raportul a fost salvat:\n{file_path}",
                                bg_color=self.frame_bg_color, fg_color=self.text_color)
            raport_window.destroy()



        tk.Button(raport_window, text="Generează și exportă Excel",
                  bg=self.accent_color, fg=self.text_color, width=25,
                  command=genereaza_raport).pack(pady=20)


        def show_custom_message(parent, title, message, bg_color, fg_color):
            msg_win = tk.Toplevel(parent)
            msg_win.title(title)
            msg_win.configure(bg=bg_color)
            msg_win.resizable(False, False)

            # Poziționare centrată față de parent
            parent.update_idletasks()
            width, height = 300, 150
            x = parent.winfo_rootx() + (parent.winfo_width() // 2) - (width // 2)
            y = parent.winfo_rooty() + (parent.winfo_height() // 2) - (height // 2)
            msg_win.geometry(f"{width}x{height}+{x}+{y}")

            # Mesaj
            tk.Label(msg_win, text=message, bg=bg_color, fg=fg_color, font=("Arial", 10, "bold"), wraplength=280).pack(
                expand=True, pady=20)

            # Buton OK
            tk.Button(msg_win, text="OK", width=10, bg=self.accent_color, fg="white", command=msg_win.destroy).pack(
                pady=5)

            msg_win.grab_set()
            msg_win.focus_set()
            msg_win.transient(parent)
            msg_win.wait_window()

